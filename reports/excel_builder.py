import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
CHEAPEST_FILL = PatternFill("solid", fgColor="2DD4BF")
STAR_FILLS    = {
    "3★": PatternFill("solid", fgColor="D6E4F0"),
    "4★": PatternFill("solid", fgColor="D5F5E3"),
    "5★": PatternFill("solid", fgColor="FEF9E7"),
    "7★": PatternFill("solid", fgColor="FDEDEC"),
}
ALT_FILL    = PatternFill("solid", fgColor="F8F9FA")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Calibri", size=10)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)
PRICE_FMT   = '£#,##0.00'
DATE_FMT    = 'D-MMM-YY'

thin   = Side(border_style="thin", color="DEE2E6")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

BOARD_ORDER = ["RO", "BB", "HB", "FB", "AI"]
BOARD_LABELS = {
    "RO": "Room Only",
    "BB": "Bed & Breakfast",
    "HB": "Half Board",
    "FB": "Full Board",
    "AI": "All Inclusive",
}


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_hotel_grid(ws, df: pd.DataFrame, title: str, star_fill=None):
    """
    Writes a grid: rows=dates, col groups=hotel+board_basis.
    Each hotel gets columns for each board type it has.
    """
    ws.title = title[:31]

    if df.empty:
        ws["A1"] = "No data for this category."
        return

    df = df.copy()
    df["check_in"] = pd.to_datetime(df["check_in"])

    # Build column structure: (hotel_name, board_basis)
    hotels     = sorted(df["hotel_name"].unique())
    boards     = [b for b in BOARD_ORDER if b in df["board_basis"].unique()]
    dates      = sorted(df["check_in"].unique())
    durations  = sorted(df["nights"].unique())

    # ── Header row 1: night durations spanning columns ───────────────────────
    col = 2
    ws.cell(row=1, column=1, value="Date").font   = HEADER_FONT
    ws.cell(row=1, column=1).fill   = HEADER_FILL
    ws.cell(row=1, column=1).border = BORDER

    for nights in durations:
        label = f"{int(nights)} Nights"
        span  = len(hotels) * len(boards)
        ws.cell(row=1, column=col, value=label).font  = HEADER_FONT
        ws.cell(row=1, column=col).fill   = HEADER_FILL
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        col += span

    # ── Header row 2: hotel names spanning board columns ─────────────────────
    col = 2
    ws.cell(row=2, column=1, value="").fill = HEADER_FILL
    for nights in durations:
        for hotel in hotels:
            ws.cell(row=2, column=col, value=hotel[:22]).font  = BOLD_FONT
            ws.cell(row=2, column=col).fill   = star_fill or ALT_FILL
            ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", wrap_text=True)
            if len(boards) > 1:
                ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + len(boards) - 1)
            col += len(boards)

    # ── Header row 3: board basis ─────────────────────────────────────────────
    col = 2
    ws.cell(row=3, column=1, value="").fill = HEADER_FILL
    for nights in durations:
        for hotel in hotels:
            for board in boards:
                cell = ws.cell(row=3, column=col, value=BOARD_LABELS.get(board, board))
                cell.font   = Font(name="Calibri", bold=True, size=9, color="444444")
                cell.fill   = HEADER_FILL
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")
                col += 1

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, dt in enumerate(dates, start=4):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        date_cell = ws.cell(row=row_idx, column=1, value=pd.Timestamp(dt).date())
        date_cell.number_format = DATE_FMT
        date_cell.font   = BODY_FONT
        date_cell.fill   = fill
        date_cell.border = BORDER

        col = 2
        for nights in durations:
            for hotel in hotels:
                sub = df[
                    (df["check_in"] == dt) &
                    (df["nights"]   == nights) &
                    (df["hotel_name"] == hotel)
                ]
                for board in boards:
                    cell = ws.cell(row=row_idx, column=col)
                    row_data = sub[sub["board_basis"] == board]
                    if not row_data.empty:
                        price = row_data["price_gbp"].min()
                        cell.value         = round(price, 2)
                        cell.number_format = PRICE_FMT
                        cell.font          = BODY_FONT
                    cell.fill   = fill
                    cell.border = BORDER
                    cell.alignment = Alignment(horizontal="right")
                    col += 1

    # ── Highlight cheapest in each row ────────────────────────────────────────
    for row in ws.iter_rows(min_row=4, min_col=2):
        prices = [(c.column, c.value) for c in row if isinstance(c.value, (int, float))]
        if not prices:
            continue
        min_val = min(v for _, v in prices)
        for col_num, val in prices:
            if val == min_val:
                ws.cell(row=row[0].row, column=col_num).fill = CHEAPEST_FILL
                ws.cell(row=row[0].row, column=col_num).font = BOLD_FONT

    # ── Column widths ─────────────────────────────────────────────────────────
    _col_width(ws, 1, 13)
    for c in range(2, col):
        _col_width(ws, c, 12)

    ws.freeze_panes = "B4"
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 22


def _write_cheapest_by_board(ws, df: pd.DataFrame, board: str, title: str):
    """Cheapest hotel per date per night duration for one board type."""
    ws.title = title[:31]
    sub = df[df["board_basis"] == board]
    if sub.empty:
        ws["A1"] = f"No {BOARD_LABELS.get(board, board)} data available."
        return

    sub = sub.copy()
    sub["check_in"] = pd.to_datetime(sub["check_in"])

    durations = sorted(sub["nights"].unique())

    # Header
    ws.cell(row=1, column=1, value="Date").fill = HEADER_FILL
    ws.cell(row=1, column=1).font   = HEADER_FONT
    ws.cell(row=1, column=1).border = BORDER
    for col, nights in enumerate(durations, start=2):
        cell = ws.cell(row=1, column=col, value=f"{int(nights)} Nights — {BOARD_LABELS.get(board, board)}")
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data
    dates = sorted(sub["check_in"].unique())
    for row_idx, dt in enumerate(dates, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        date_cell = ws.cell(row=row_idx, column=1, value=pd.Timestamp(dt).date())
        date_cell.number_format = DATE_FMT
        date_cell.font   = BODY_FONT
        date_cell.fill   = fill
        date_cell.border = BORDER

        for col, nights in enumerate(durations, start=2):
            cell = ws.cell(row=row_idx, column=col)
            row_data = sub[(sub["check_in"] == dt) & (sub["nights"] == nights)]
            if not row_data.empty:
                best_idx  = row_data["price_gbp"].idxmin()
                best      = row_data.loc[best_idx]
                cell.value         = round(best["price_gbp"], 2)
                cell.number_format = PRICE_FMT
                cell.font          = BODY_FONT
                cell.comment_text  = best.get("hotel_name", "")
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right")

    _col_width(ws, 1, 13)
    for c in range(2, len(durations) + 2):
        _col_width(ws, c, 18)
    ws.freeze_panes = "B2"


def _write_csv_sheet(ws, df: pd.DataFrame):
    ws.title = "CSV"
    if df.empty:
        ws["A1"] = "No data"
        return
    cols = ["hotel_name", "stars", "check_in", "check_out", "nights",
            "board_basis", "board_label", "price_gbp", "source"]
    available = [c for c in cols if c in df.columns]
    ws.append(available)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for _, row in df[available].iterrows():
        ws.append([str(v) if isinstance(v, pd.Timestamp) else v for v in row.tolist()])


def build_excel(
    hotels_df: pd.DataFrame,
    flights_df: pd.DataFrame,
    packages_df: pd.DataFrame,
    departure_cities: list,
    night_durations: list,
    output_dir: str,
    destination: str = "AMS",
) -> str:

    wb = Workbook()
    wb.remove(wb.active)

    if hotels_df.empty:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No hotel data was returned. Check credentials and destination code."
    else:
        df = hotels_df.copy()
        df["check_in"] = pd.to_datetime(df["check_in"])

        # ── All Hotels tab ────────────────────────────────────────────────────
        ws_all = wb.create_sheet("All Hotels")
        _write_hotel_grid(ws_all, df, "All Hotels")

        # ── Per star-rating tabs ──────────────────────────────────────────────
        available_stars = sorted(
            [s for s in df["stars"].unique() if s != "N/A"],
            key=lambda x: int(x.replace("★", "")),
            reverse=True,
        )
        for star in available_stars:
            sub = df[df["stars"] == star]
            tab_name = f"{star} Hotels"
            ws_star  = wb.create_sheet(tab_name)
            fill     = STAR_FILLS.get(star)
            _write_hotel_grid(ws_star, sub, tab_name, star_fill=fill)

        # ── Cheapest per board type ───────────────────────────────────────────
        boards_present = [b for b in BOARD_ORDER if b in df["board_basis"].unique()]
        for board in boards_present:
            label    = BOARD_LABELS.get(board, board)
            ws_board = wb.create_sheet(f"Best {board}")
            _write_cheapest_by_board(ws_board, df, board, f"Best {label}")

        # ── Flights tab ───────────────────────────────────────────────────────
        if not flights_df.empty:
            from transform.engine import flight_price_pivot
            ws_fl = wb.create_sheet("Flights")
            for city in departure_cities[:1]:
                p = flight_price_pivot(flights_df, city, night_durations)
                if not p.empty:
                    _write_cheapest_by_board.__doc__  # just a no-op
                    ws_fl["A1"] = f"Flights from {city} → {destination}"

        # ── CSV raw export ────────────────────────────────────────────────────
        ws_csv = wb.create_sheet("CSV")
        _write_csv_sheet(ws_csv, df)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename  = f"{destination}_hotels_{timestamp}.xlsx"
    filepath  = os.path.join(output_dir, filename)
    wb.save(filepath)

    # Also save as JSON for the dashboard
    if not hotels_df.empty:
        json_path = os.path.join(output_dir, "latest.json")
        hotels_df.copy().assign(
            check_in=hotels_df["check_in"].astype(str),
            check_out=hotels_df["check_out"].astype(str),
        ).to_json(json_path, orient="records", indent=2)

    return filepath
